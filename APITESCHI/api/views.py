from django.shortcuts import render,redirect
from rest_framework.views import APIView
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.core.mail import send_mail
from django.template.loader import render_to_string

# Cadena aleatoria
import secrets
import string

from django.utils.datastructures import MultiValueDictKeyError
from django.shortcuts import render
#Exportar a excel
from django.http import HttpResponse
import openpyxl
# Importa el modelo de usuarios de Django
from django.contrib.auth.models import User 
# Create your views here.


class Home(APIView):
    template_name="index.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name)

class Signup(APIView):
    template_name="signup.html"
    def get(self, request):
        return render(request, self.template_name,{
            'form' : UserCreationForm
        })
    def post(self, request):
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(first_name=request.POST['first_name'],last_name=request.POST['last_name'], username=request.POST['username'], password=request.POST['password1'], email=request.POST['email'])
                user.save()
                correo = request.POST['email']
                usuario = request.POST['username']
                contra = request.POST['password1']
                asunto = 'Registro'
                return redirect('enviar_correo', correo=correo, usuario=usuario, contra=contra, asunto=asunto)
                
            except IntegrityError:
                return render(request, self.template_name,{
                    'form' : UserCreationForm,
                    "mensaje" : 'Username already exist'
                })
                #return HttpResponse('Username already exist')
        return render(request, self.template_name,{
                    'form' : UserCreationForm,
                    "mensaje" : 'Password do not match'
                })

class Signout(APIView):
    def get(self, request):
        logout(request)
        return redirect('signin')
    def post(self, request):
        logout(request)
        return redirect('signin')  

class Signin(APIView):
    template_name="signin.html"
    def get(self, request):
        return render(request, self.template_name, {
            'form': AuthenticationForm
        })
    def post(self, request):
        try:
            user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
            if user is None:
                return render(request, 'signin.html', {
                    'form': AuthenticationForm,
                    'error': 'Usuario o contraseña incorrecta'
                })
            else:
                login(request, user)
                return redirect ('/')
        except MultiValueDictKeyError:
            return render(request, self.template_name, {
                    'form': AuthenticationForm,
                    'error': 'Alguno de los campos no han sido llenados de manera correcta, intentalo de nuevo'
                })
        except IntegrityError:
            return render(request, self.template_name, {
                'form' : UserCreationForm,
                "mensaje" : 'No se ha podido iniciar sesión de manera coorrecta, intentalo de nuevo'
            })

class About(APIView):
    template_name="about.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name)

class Contact(APIView):
    template_name="contact.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name)   

class ShopS(APIView):
    template_name="shop-single.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name) 
    
class Shop(APIView):
    template_name="shop.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name)

class Forgot(APIView):
    template_name="forgot.html"
    def get(self,request):
        return render(request, self.template_name,{
            'form' : UserCreationForm
        })
    def post(self,request):
        longitud = 10  # Longitud de la contraseña
        caracteres = string.ascii_letters + string.digits  # Caracteres permitidos
        contra_aleatoria = ''.join(secrets.choice(caracteres) for _ in range(longitud))
        try:
            user = User.objects.filter(email=request.POST['email'])
            if user.exists():
                user = user[0]
                usuario = user.username
                user.set_password(contra_aleatoria)
                user.save()
                # Defines variables para que posteriormente las mandes por una mamada de link inverso xd a la clase que manda el correo
                asunto = "Recuperar contraseña"
                correo = request.POST['email']
                usuario = usuario
                # Igual aqui a contra le mandas el valor de la cadena generada automaticamente
                contra = contra_aleatoria
                # Aqui retorna a la clase de enviar correo
                return redirect('enviar_correo', correo=correo, usuario=usuario, contra=contra, asunto=asunto)
            else:
                return render(request, self.template_name,{
                    'form' : UserCreationForm,
                    "mensaje" : 'No hemos podido localizar tu cuenta, asegurate de que tu correo sea correcto'
                })
        except IntegrityError:
            return render(request, self.template_name,{
                'form' : UserCreationForm,
                "mensaje" : 'No hemos podido localizar tu cuenta, asegurate de que tu correo sea correcto'
            })
        
class Table(APIView):
    template_name="table.html"
    def get(self, request):
        usuarios = User.objects.all()  # Obtén todos los registros de la tabla auth_user
        return render(request, 'table.html', {'usuarios': usuarios})
    def post(self, request):
        return render(request,self.template_name)    

class Chart(APIView):
    template_name="chartjs.html"
    def get(self, request):
         return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name) 
"""    
class Signin(APIView):
    template_name="signin.html"
    def get(self, request):
        return render(request,self.template_name)
    def post(self, request):
        return render(request,self.template_name)
"""
#class Signup(APIView):
#    template_name="signup.html"
#    def get(self, request):
#        return render(request,self.template_name)
#    def post(self, request):
#        return render(request,self.template_name)

# envio de correo

def enviar_correo(request, correo, usuario, contra, asunto):
    subject = asunto
    from_email = 'jhonygenial12@gmail.com'
    recipient_list = [correo]

    # Renderiza la plantilla HTML con el contexto
    contexto = {'asunto': asunto,
                'usuario': usuario,
                "contra": contra}
    contenido_correo = render_to_string('sendcorreo.html', contexto)

    # Envía el correo
    send_mail(subject, '', from_email, recipient_list, html_message=contenido_correo)
    
    return redirect('signin')

def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Agrega encabezados de columna
    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Nombre de usuario'
    worksheet['C1'] = 'Correo electrónico'

    # Obtiene los datos de la tabla auht_users
    users = User.objects.all()

    # Llena el archivo Excel con los datos de la tabla
    row_num = 2
    for user in users:
        worksheet.cell(row=row_num, column=1, value=user.id)
        worksheet.cell(row=row_num, column=2, value=user.username)
        worksheet.cell(row=row_num, column=3, value=user.email)
        row_num += 1

    workbook.save(response)

    return response